"""
Convert all Excel files in input/ to JSON files in output/.
Output filenames mirror the input directory structure, e.g.:
  input/semgrep/C-semgrep-merged.xlsx -> output/semgrep/C-semgrep-merged.json
"""

import json
import os
from pathlib import Path

import openpyxl

INPUT_DIR = Path("input")
OUTPUT_DIR = Path("output")


def sheet_to_records(ws) -> list[dict]:
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        records.append(dict(zip(headers, row)))
    return records


def convert(xlsx_path: Path) -> Path:
    rel = xlsx_path.relative_to(INPUT_DIR)
    out_path = OUTPUT_DIR / rel.with_suffix(".json")
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    if len(wb.sheetnames) == 1:
        data = sheet_to_records(wb.active)
    else:
        data = {name: sheet_to_records(wb[name]) for name in wb.sheetnames}

    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2, default=str)

    wb.close()
    return out_path


def main():
    xlsx_files = sorted(INPUT_DIR.rglob("*.xlsx"))
    if not xlsx_files:
        print("No .xlsx files found in input/")
        return

    for xlsx_path in xlsx_files:
        out_path = convert(xlsx_path)
        print(f"  {xlsx_path}  ->  {out_path}")

    print(f"\nDone. {len(xlsx_files)} file(s) converted.")


if __name__ == "__main__":
    main()
